import openpyxl
import json
from django.views.decorators.http import require_POST
from .models import Producto

from django.db import models

from .models import Producto, CategoriaProducto, Bodega


from io import BytesIO
from django.utils.timezone import now
from django.http import HttpResponse
from django.utils import timezone

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.http import require_POST
from datetime import date, timedelta

from .models import Producto, HistorialPrecio, MovimientoInventario, CategoriaProducto
from .forms import ProductoForm, UsuarioForm, MovimientoInventarioForm

# 🎯 Decorador: permite un grupo o admin
def group_required_or_admin(group_name):
    def check_group(user):
        return user.is_authenticated and (
            user.groups.filter(name=group_name).exists() or user.groups.filter(name='Administrador').exists()
        )
    return user_passes_test(check_group)

# ✅ Decorador: permite múltiples grupos o admin
def groups_required(*group_names):
    def check_groups(user):
        return user.is_authenticated and (
            user.groups.filter(name__in=group_names).exists() or user.groups.filter(name='Administrador').exists()
        )
    return user_passes_test(check_groups)

def index(request):
    return render(request, 'inventario/index.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            if user.groups.filter(name='Administrador').exists():
                return redirect('usuarios_crud')
            elif user.groups.filter(name='Gestor de Inventario').exists():
                return redirect('producto_crud')
            elif user.groups.filter(name='Comprador').exists():
                return redirect('tienda')
            else:
                return redirect('index')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'inventario/login.html')

@login_required
@groups_required('Comprador', 'Gestor de Inventario', 'Almacén')
def lista_productos_view(request):
    productos = Producto.objects.all()
    return render(request, 'inventario/lista_productos.html', {'productos': productos})

def registro_usuario(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')

        if password != password_confirm:
            return render(request, 'inventario/registro_usuario.html', {
                'error': "Las contraseñas no coinciden."
            })

        if username and email and password:
            if not User.objects.filter(username=username).exists():
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=password
                )
                # 🔥 No se asigna grupo aquí, queda sin rol
                return redirect('login')
            else:
                return render(request, 'inventario/registro_usuario.html', {
                    'error': "El nombre de usuario ya existe."
                })
    return render(request, 'inventario/registro_usuario.html')


def olvidar_contra(request):
    return render(request, 'inventario/olvidar_contra.html')

def registro_inventario(request):
    return render(request, 'inventario/registro_item.html')

@login_required
@groups_required('Gestor de Inventario', 'Almacén')
def producto_crud(request):
    categoria_id = request.GET.get('categoria')
    bodega_id = request.GET.get('bodega')  # 👈 NUEVO

    categorias = CategoriaProducto.objects.order_by('nombre')
    bodegas = Bodega.objects.order_by('nombre')  # 👈 NUEVO
    productos = Producto.objects.all()

    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)

    if bodega_id:
        productos = productos.filter(bodega_id=bodega_id)  # 👈 NUEVO

    if 'delete' in request.GET:
        producto = get_object_or_404(Producto, pk=request.GET['delete'])
        producto.delete()
        return redirect('producto_crud')

    return render(request, 'inventario/producto_crud.html', {
        'productos': productos,
        'categorias': categorias,
        'bodegas': bodegas,  # 👈 NUEVO
        'categoria_seleccionada': int(categoria_id) if categoria_id else None,
        'bodega_seleccionada': int(bodega_id) if bodega_id else None,
        'today': date.today(),
        'today_plus_7': date.today() + timedelta(days=7),
    })

@login_required
@group_required_or_admin('Gestor de Inventario')
def registro_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            rol = request.user.groups.first().name if request.user.groups.exists() else "Sin rol"
            HistorialPrecio.objects.create(
                producto=producto,
                precio=producto.precio,
                usuario=request.user,
                rol=rol
            )
            return redirect('producto_crud')
    else:
        form = ProductoForm()
    return render(request, 'inventario/registro_producto.html', {'form': form})

@login_required
@groups_required('Gestor de Inventario', 'Almacén')
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    precio_anterior = producto.precio
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            producto_actualizado = form.save()
            if producto_actualizado.precio != precio_anterior:
                rol = request.user.groups.first().name if request.user.groups.exists() else "Sin rol"
                HistorialPrecio.objects.create(
                    producto=producto_actualizado,
                    precio=producto_actualizado.precio,
                    usuario=request.user,
                    rol=rol
                )
            messages.success(request, "Producto modificado exitosamente.")
            return redirect('producto_crud')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'inventario/registro_producto.html', {
        'form': form,
        'editar': True,
        'producto': producto
    })

@login_required
@group_required_or_admin('Administrador')
def usuarios_crud(request):
    editar = False
    usuario = None
    grupo_usuario = None

    if 'edit' in request.GET:
        editar = True
        usuario = get_object_or_404(User, pk=request.GET['edit'])
        form = UsuarioForm(request.POST or None, instance=usuario)
        grupo_usuario = usuario.groups.first()
    elif 'delete' in request.GET:
        usuario = get_object_or_404(User, pk=request.GET['delete'])
        usuario.delete()
        messages.success(request, "Usuario eliminado exitosamente.")
        return redirect('usuarios_crud')
    else:
        form = UsuarioForm(request.POST or None)

    if request.method == 'POST':
        if editar:
            form = UsuarioForm(request.POST, instance=usuario)
        else:
            form = UsuarioForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            if form.cleaned_data['password']:
                user.set_password(form.cleaned_data['password'])
            user.save()
            selected_group_name = request.POST.get('grupo')
            if selected_group_name:
                group = Group.objects.filter(name=selected_group_name).first()
                user.groups.clear()
                if group:
                    user.groups.add(group)
            messages.success(request, "Usuario guardado exitosamente.")
            return redirect('usuarios_crud')

    grupo_seleccionado = request.GET.get('grupo')
    usuarios = User.objects.all()
    if grupo_seleccionado:
        usuarios = usuarios.filter(groups__name=grupo_seleccionado)

    grupos = Group.objects.all()
    return render(request, 'inventario/usuarios_crud.html', {
        'form': form,
        'usuarios': usuarios,
        'editar': editar,
        'usuario': usuario,
        'grupos': grupos,
        'grupo_usuario': grupo_usuario,
        'grupo_seleccionado': grupo_seleccionado,
    })

@login_required
@group_required_or_admin('Administrador')
def historial_precios_producto(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    historial = producto.historial_precios.all().order_by('-fecha')
    return render(request, 'inventario/historial_precios.html', {
        'producto': producto,
        'historial': historial
    })

@login_required
@group_required_or_admin('Administrador')
def eliminar_historial_precio(request, historial_id):
    historial = get_object_or_404(HistorialPrecio, pk=historial_id)
    historial.delete()
    messages.success(request, "Historial eliminado.")
    return redirect('producto_crud')

@login_required
@group_required_or_admin('Administrador')
def eliminar_movimiento(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, pk=movimiento_id)
    movimiento.delete()
    messages.success(request, "Movimiento eliminado.")
    return redirect('producto_crud')

@login_required
@groups_required('Gestor de Inventario', 'Almacén')
def registrar_movimiento(request):
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.usuario = request.user
            movimiento.rol = request.user.groups.first().name if request.user.groups.exists() else "Sin rol"
            movimiento.save()
            if movimiento.tipo == 'entrada':
                movimiento.producto.stock += movimiento.cantidad
            elif movimiento.tipo == 'salida':
                movimiento.producto.stock -= movimiento.cantidad
            movimiento.producto.save()
            messages.success(request, "Movimiento registrado correctamente.")
            return redirect('listar_movimientos')
    else:
        form = MovimientoInventarioForm()
    return render(request, 'inventario/registrar_movimiento.html', {'form': form})

@login_required
@groups_required('Gestor de Inventario', 'Almacén')
def listar_movimientos(request):
    movimientos = MovimientoInventario.objects.all().order_by('-fecha')
    return render(request, 'inventario/listar_movimientos.html', {'movimientos': movimientos})

@login_required
@groups_required('Gestor de Inventario', 'Almacén')
@require_POST
def registrar_movimiento_directo(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    tipo = request.POST.get('tipo')
    cantidad = int(request.POST.get('cantidad', 0))
    motivo = request.POST.get('motivo', '')

    if tipo in ['entrada', 'salida'] and cantidad > 0:
        movimiento = MovimientoInventario.objects.create(
            producto=producto,
            tipo=tipo,
            cantidad=cantidad,
            motivo=motivo,
            usuario=request.user,
            rol=request.user.groups.first().name if request.user.groups.exists() else "Sin rol"
        )
        if tipo == 'entrada':
            producto.stock += cantidad
        else:
            producto.stock = max(producto.stock - cantidad, 0)
        producto.save()

        messages.success(request, f"Movimiento '{tipo}' registrado para {producto.nombre}.")
    else:
        messages.error(request, "Error al registrar movimiento.")
    return redirect('producto_crud')

def ver_carrito(request):
    return render(request, 'inventario/carrito.html')

def logout_view(request):
    logout(request)
    return redirect('index')

@login_required
@groups_required('Administrador', 'Gestor de Inventario', 'Almacén','Comprador')
def panel_control_view(request):
    total_productos = Producto.objects.count()
    stock_critico = Producto.objects.filter(stock__lt=5).count()
    total_usuarios = User.objects.count()
    usuarios_activos = User.objects.filter(last_login__date=timezone.now().date()).count()

    categorias = CategoriaProducto.objects.all()
    labels_categorias = [c.nombre for c in categorias]
    data_categorias = [Producto.objects.filter(categoria=c).count() for c in categorias]

    productos = Producto.objects.all()
    labels_productos = [p.nombre for p in productos[:6]]
    data_stock = [p.stock for p in productos[:6]]

    # 🔔 Productos sin stock o en stock mínimo
    productos_stock_minimo = Producto.objects.filter(stock__lte=models.F('stock_minimo'))

    context = {
        'total_productos': total_productos,
        'stock_critico': stock_critico,
        'total_usuarios': total_usuarios,
        'usuarios_activos': usuarios_activos,
        'labels_categorias': json.dumps(labels_categorias),
        'data_categorias': json.dumps(data_categorias),
        'labels_productos': json.dumps(labels_productos),
        'data_stock': json.dumps(data_stock),
        'productos_stock_minimo': productos_stock_minimo,  # 📌
    }

    return render(request, 'inventario/panel_control.html', context)

@login_required
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws_productos = wb.active
    ws_productos.title = "Productos"

    ws_productos.append(['ID', 'Nombre', 'Categoría', 'Precio', 'Stock'])

    for p in Producto.objects.all():
        ws_productos.append([
            p.id,
            p.nombre,
            p.categoria.nombre if p.categoria else '',
            float(p.precio),
            p.stock
        ])

    # Estadísticas
    ws_stats = wb.create_sheet(title="Estadísticas")
    ws_stats.append(["Métrica", "Valor"])
    ws_stats.append(["Total Productos", Producto.objects.count()])
    ws_stats.append(["Stock Crítico (<=5)", Producto.objects.filter(stock__lte=5).count()])
    ws_stats.append(["Usuarios Registrados", User.objects.count()])
    ws_stats.append(["Usuarios Activos Hoy", User.objects.filter(last_login__date=date.today()).count()])

    # Movimientos de inventario
    ws_mov = wb.create_sheet(title="Movimientos")
    ws_mov.append(['ID', 'Producto', 'Tipo', 'Cantidad', 'Usuario', 'Fecha'])

    for mov in MovimientoInventario.objects.all():
        ws_mov.append([
            mov.id,
            mov.producto.nombre,
            mov.tipo,
            mov.cantidad,
            mov.usuario.username,
            mov.fecha.strftime('%Y-%m-%d')
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_maestranza.xlsx'
    return response

    
@login_required
@groups_required('Administrador', 'Gestor de Inventario')
@require_POST
def cambiar_bodega(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    nueva_bodega_id = request.POST.get('nueva_bodega')

    try:
        nueva_bodega = Bodega.objects.get(id=nueva_bodega_id)
        producto.bodega = nueva_bodega
        producto.save()
        messages.success(request, f"Bodega del producto '{producto.nombre}' actualizada a '{nueva_bodega.nombre}'.")
    except Bodega.DoesNotExist:
        messages.error(request, "La bodega seleccionada no existe.")

    return redirect('producto_crud')

